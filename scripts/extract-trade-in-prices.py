#!/usr/bin/env python3
"""Extract Atlas Mobile buyback prices (iPhone / Samsung / iPad) from the
uploaded .xlsx into a normalised JSON the storefront can drive the trade-in
flow from. Re-run whenever a fresh price book is provided.

Usage: python3 scripts/extract-trade-in-prices.py <file.xlsx> <out.json>
"""
import json, re, sys
import openpyxl

def money(v):
    if v is None: return None
    s = str(v).strip()
    if s in ("", "#NUM!", "#REF!", "#VALUE!"): return None
    if re.search(r'not buying|ask|n/?a', s, re.I): return None
    m = re.search(r'-?\d+(?:\.\d+)?', s.replace(",", ""))
    return int(round(float(m.group(0)))) if m else None

def gb(s):
    m = re.search(r'(\d+(?:\.\d+)?)\s*(TB|GB)\b', s, re.I)
    if not m: return 0
    n = float(m.group(1))
    return int(round(n*1024)) if m.group(2).upper()=="TB" else int(round(n))

def lock_of(s):
    u = s.lower()
    if "unlocked" in u: return "unlocked"
    if "at&t" in u or "att" in u: return "att"
    if "carrier locked" in u or "locked" in u: return "locked"
    return None

def parse_ded(note):
    """From an iPhone group note → {crackedBack, crackedLens, faceId}."""
    d = {}
    if not note: return d
    m = re.search(r'Cracked Back\s*=\s*\$?(\d+)', note, re.I)
    if m: d["crackedBack"] = int(m.group(1))
    m = re.search(r'Cracked Lens\s*=\s*\$?(\d+)', note, re.I)
    if m: d["crackedLens"] = int(m.group(1))
    m = re.search(r'Bad Face ID\s*=\s*([A-Za-z ]+?)(?:\s*/|$)', note, re.I)
    if m: d["faceId"] = m.group(1).strip()
    return d

F, OUT = sys.argv[1], sys.argv[2]
wb = openpyxl.load_workbook(F, data_only=True)
rows = []

# ---------- iPhone Used : B=Model C=SWAP D=A E=B F=C G=D H=DOA -------------
ws = wb['iPhone Used']
cur_ded = {}
for r in range(1, ws.max_row+1):
    b = ws.cell(r,2).value
    if b is None: continue
    s = str(b).strip()
    if 'Cracked Back' in s:
        cur_ded = parse_ded(s); continue
    if s.lower() == 'model': continue
    if not re.search(r'\d+\s*(GB|TB)', s, re.I): continue
    lock = lock_of(s)
    if lock is None: continue
    base = re.sub(r'\s*\d+\s*(GB|TB).*$', '', s, flags=re.I).strip()
    entry = {
        "cat":"iphone","model":base,"gb":gb(s),"lock":lock,
        "swap":money(ws.cell(r,3).value),"a":money(ws.cell(r,4).value),
        "b":money(ws.cell(r,5).value),"c":money(ws.cell(r,6).value),
        "d":money(ws.cell(r,7).value),"doa":money(ws.cell(r,8).value),
    }
    entry.update(cur_ded)
    rows.append(entry)

# ---------- Samsung : header D=NEW E=A F=B G=C H=D I=DOA ; C=variant --------
ws = wb['Samsung']
cur_model = None
for r in range(1, ws.max_row+1):
    b = ws.cell(r,2).value; c = ws.cell(r,3).value
    bs = str(b).strip() if b else ""
    cs = str(c).strip() if c else ""
    if bs.lower().startswith("galaxy"):
        cur_model = bs
    lock = lock_of(cs)
    if lock is None or cur_model is None: continue
    black = 0
    m = re.search(r'Black\s*-\$?(\d+)', cs, re.I)
    if m: black = int(m.group(1))
    rows.append({
        "cat":"samsung","model":cur_model,"gb":0,"lock":lock,
        "new":money(ws.cell(r,4).value),"a":money(ws.cell(r,5).value),
        "b":money(ws.cell(r,6).value),"c":money(ws.cell(r,7).value),
        "d":money(ws.cell(r,8).value),"doa":money(ws.cell(r,9).value),
        **({"blackDed":black} if black else {}),
    })

# ---------- iPad Used : B=Model C=A D=B E=C F=D G=DOA ----------------------
ws = wb['iPad Used']
for r in range(1, ws.max_row+1):
    b = ws.cell(r,2).value
    if b is None: continue
    s = str(b).strip()
    if s.lower().startswith('grade') or 'Grade A' in str(ws.cell(r,3).value or ''):
        continue
    if not re.search(r'\d+\s*(GB|TB)', s, re.I): continue
    base = re.sub(r'\s*\d+\s*(GB|TB).*$', '', s, flags=re.I).strip()
    rows.append({
        "cat":"ipad","model":base,"gb":gb(s),"lock":"unlocked",
        "a":money(ws.cell(r,3).value),"b":money(ws.cell(r,4).value),
        "c":money(ws.cell(r,5).value),"d":money(ws.cell(r,6).value),
        "doa":money(ws.cell(r,7).value),
    })

json.dump(rows, open(OUT,"w"), indent=0)
print(f"wrote {len(rows)} rows -> {OUT}")
from collections import Counter
print("by cat:", Counter(r["cat"] for r in rows))
# spot checks
for want in ["iPhone 17 Pro Max","iPhone 13","iPhone 7"]:
    ex=[r for r in rows if r["cat"]=="iphone" and r["model"]==want and r["gb"] and r["lock"]=="unlocked"]
    if ex: print("  CHK", want, ex[0])
for want in ["Galaxy S24 Ultra","Galaxy S24"]:
    ex=[r for r in rows if r["cat"]=="samsung" and r["model"]==want]
    if ex: print("  CHK", want, ex[0])
